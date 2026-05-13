#!/usr/bin/env python3
"""Generate discovery-phase-completion-merrakii.xlsx from hard-coded table mirror of deliverable HTML."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "deliverable" / "discovery-phase-completion-merrakii.xlsx"

GROUPS: list[tuple[str, str, str, str, list[tuple[str, str]]]] = [
    (
        "1",
        "",
        "Project kick-off & governance",
        "Measurable: operating model for the engagement is on record",
        [
            (
                "Formal kick-off session held with Merrakii and Maple AI Technologies.",
                "Kick-off session completed on 22nd April 2026.",
            ),
            (
                "Recurring review cadence defined (frequency, attendees, purpose).",
                "Weekly cadence set up to meet at Merrakii office on Wednesday and Friday.",
            ),
            (
                "Project POCs identified on client (Merrakii) and vendor (Maple AI Technologies) sides.",
                "Ruchii Suri — POC on Merrakii side; Tarun Rastogi — POC on Maple AI Technologies side.",
            ),
        ],
    ),
    (
        "2",
        "Category A",
        "Brand presentation & business clarity",
        "Website colour theme captured; business purpose captured; business intent captured.",
        [
            (
                "Website colour theme (primary, secondary, accent and usage notes) reviewed with Merrakii and accepted as the baseline for design.",
                "Approved colour reference captured for UI implementation and component styling.",
            ),
            (
                "Business purpose for the Merrakii website articulated by stakeholders and confirmed as accurate.",
                "Signed-off business purpose statement stored in the discovery pack.",
            ),
            (
                "Business intent (what Merrakii expects the website to achieve) documented and agreed with stakeholders.",
                "Business intent baseline referenced for UX, content priority and success tracking.",
            ),
        ],
    ),
    (
        "3",
        "Category A",
        "Site pages, services & functional categories",
        "Pages of the website; services to be offered; functional categories — all identified and agreed.",
        [
            (
                "Required pages for launch (and purpose of each) listed and mutually agreed with Merrakii.",
                "Agreed page list baselined for information architecture and development scope.",
            ),
            (
                "Services Merrakii will present on the website enumerated and validated with stakeholders.",
                "Services baseline locked for content blocks and catalogue-style UX.",
            ),
            (
                "Functional categories for the product experience (e.g. discovery, applications, accounts, payments) named and agreed.",
                "Functional category map aligned to roadmap modules and build sequencing.",
            ),
        ],
    ),
    (
        "4",
        "Category A",
        "Client-authored website content",
        "Website content the client must provide — collected (requirements listed; supplied assets received or logged).",
        [
            (
                "Copy, imagery, institute/service material and other Merrakii-supplied content requirements listed; supplied items received or logged.",
                "Client content intake register completed for tracking outstanding vs received assets.",
            ),
        ],
    ),
    (
        "5",
        "Category B",
        "Integrations, infrastructure & commercial",
        "Integrations; DNS; payment gateway artefacts; roadmap (Maple AI Technologies); quotation.",
        [
            (
                "Required integrations (e.g. CRM, email capture, analytics, payments) identified with Merrakii and agreed in principle.",
                "Approved integration list referenced for technical design and vendor setup.",
            ),
            (
                "Domain registrar/DNS details needed for deployment and cutover gathered from Merrakii or agreed access path.",
                "DNS assumptions and records checklist filed for go-live execution.",
            ),
            (
                "Payment gateway vendor forms/checklists issued to Merrakii; completed artefacts received back as required for onboarding.",
                "Client-returned payment gateway pack on file for activation with the chosen provider.",
            ),
            (
                "Phased delivery roadmap (design → build → QA → launch → hypercare or equivalent) presented by Maple AI Technologies and understood by Merrakii.",
                "Roadmap artefact issued to Merrakii and treated as the baseline schedule reference pending contract milestones.",
            ),
            (
                "Commercial quotation for agreed Discovery outcomes and subsequent phases shared with Merrakii and mutually accepted.",
                "Agreed quotation version referenced for proposal/SOW or purchase order alignment.",
            ),
        ],
    ),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Discovery deliverables"

    ws.merge_cells("A1:G1")
    ws["A1"] = "Discovery Phase — Completion Summary | Merrakii"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(
        [
            "Prepared by: Maple AI Technologies",
            "",
            "",
            "",
            "Prepared for: Munjal Universal Consultancy LLP",
            "",
            "Issue date: 13 May 2026",
        ]
    )
    ws.append([])

    headers = [
        "Table row",
        "Category",
        "Deliverable",
        "Bundle / notes",
        "What was produced & acceptance note",
        "Value / outcome (per acceptance point)",
        "Status",
    ]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="E8EAEF")
    hdr_font = Font(bold=True)
    r = ws.max_row
    for c in range(1, 8):
        cell = ws.cell(r, c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    status = "Complete"

    for t_row, cat, deliverable, notes, pairs in GROUPS:
        start = ws.max_row + 1
        for i, (acc, val) in enumerate(pairs):
            ws.append(
                [
                    t_row if i == 0 else "",
                    cat if i == 0 else "",
                    deliverable if i == 0 else "",
                    notes if i == 0 else "",
                    acc,
                    val,
                    status if i == 0 else "",
                ]
            )
        end = ws.max_row
        if end > start:
            for col in (1, 2, 3, 4, 7):
                ws.merge_cells(f"{get_column_letter(col)}{start}:{get_column_letter(col)}{end}")
        for rr in range(start, end + 1):
            for cc in range(1, 8):
                ws.cell(rr, cc).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(start, 7).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    widths = (10, 14, 28, 36, 42, 42, 12)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
