"""
Serve deliverable/digi.html (and sibling static assets under deliverable/) plus files from docs/
(e.g. executive-inputs-config.js, Excel), and POST /api/executive-inputs to upsert into the
"Submission log" sheet in docs/Merrakii_Business_Capture.xlsx (one row per Section + Q#; resubmit updates).

Storage modes
-------------
* Local: write to a path on disk (``EXECUTIVE_INPUTS_XLSX`` or default ``docs/`` in repo).
* GitHub: set ``EXECUTIVE_INPUTS_GITHUB_TOKEN`` and ``EXECUTIVE_INPUTS_GITHUB_REPO`` (e.g. ``owner/name``);
  the service GETs the file, upserts log rows, PUTs a commit. Required for static GitHub Pages
  so answers are stored in the repository workbook.

Run: python executive-inputs-server/app.py
Or:  python -m executive-inputs-server.app
"""

from __future__ import annotations

import base64
import json
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import requests
from flask import Flask, jsonify, request, send_from_directory
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DOCS = ROOT / "docs"
DELIVERABLE = ROOT / "deliverable"
DEFAULT_XLSX = DOCS / "Merrakii_Business_Capture.xlsx"
SUBMIT_TOKEN = os.environ.get("EXECUTIVE_INPUTS_SUBMIT_TOKEN", "").strip()

GITHUB_TOKEN = os.environ.get("EXECUTIVE_INPUTS_GITHUB_TOKEN", "").strip()
GITHUB_REPO = os.environ.get("EXECUTIVE_INPUTS_GITHUB_REPO", "").strip()
GITHUB_FILE = os.environ.get("EXECUTIVE_INPUTS_GITHUB_FILE", "docs/Merrakii_Business_Capture.xlsx").strip(
    "/"
)
GITHUB_BRANCH = os.environ.get("EXECUTIVE_INPUTS_GITHUB_BRANCH", "main").strip()
GITHUB_API = "https://api.github.com"

app = Flask(__name__)

LOG_SHEET = "Submission log"
HEADERS = ["Timestamp (UTC)", "Section", "Q#", "Question", "Response"]


def _gh_headers() -> dict:
    return {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }


@dataclass
class FileState:
    content: bytes
    sha: str | None  # None only if 404


def github_get_file() -> FileState:
    if not GITHUB_TOKEN or not GITHUB_REPO:
        raise RuntimeError("GitHub mode requires EXECUTIVE_INPUTS_GITHUB_TOKEN and EXECUTIVE_INPUTS_GITHUB_REPO")
    url = f"{GITHUB_API}/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"
    r = requests.get(url, headers=_gh_headers(), params={"ref": GITHUB_BRANCH}, timeout=60)
    if r.status_code == 404:
        return FileState(content=b"", sha=None)
    r.raise_for_status()
    data = r.json()
    if isinstance(data, list):
        raise RuntimeError(f"Path {GITHUB_FILE} is a directory, not a file")
    b64 = data.get("content", "")
    raw = base64.b64decode(b64.replace("\n", "")) if b64 else b""
    return FileState(content=raw, sha=data.get("sha"))


def github_put_file(content: bytes, message: str, expected_sha: str | None) -> bool:
    """Return True on success, False on 409/422 so caller can retry after refresh."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        raise RuntimeError("GitHub mode requires token and repo")
    url = f"{GITHUB_API}/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"
    body = {
        "message": message,
        "content": base64.b64encode(content).decode("ascii"),
        "branch": GITHUB_BRANCH,
    }
    if expected_sha:
        body["sha"] = expected_sha
    r = requests.put(url, headers=_gh_headers(), data=json.dumps(body), timeout=60)
    if r.status_code in (200, 201):
        return True
    if r.status_code in (409, 422):
        return False
    r.raise_for_status()
    return False


def _ensure_log_sheet(wb) -> object:
    if LOG_SHEET in wb.sheetnames:
        ws = wb[LOG_SHEET]
    else:
        ws = wb.create_sheet(LOG_SHEET, 0)
    if not ws.max_row or ws.max_row < 1 or not ws.cell(1, 1).value:
        for i, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=i, value=h)
    elif str(ws.cell(1, 1).value).strip() != HEADERS[0]:
        for i, h in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=i, value=h)
    return ws


def _question_data_rows(ws):
    """1-based data rows: column A = question (after header and optional intro row with '—')."""
    max_r = ws.max_row or 1
    intro_offset = 0
    v2 = ws.cell(row=2, column=1).value
    if v2 is not None and str(v2).strip() == "—":
        intro_offset = 1
    out = []
    for r in range(2 + intro_offset, max_r + 1):
        ac = ws.cell(row=r, column=1).value
        if ac is None:
            continue
        t = str(ac).strip()
        if not t or t == "—":
            continue
        out.append(r)
    return out


def _coerce_qnum_1based(cell_val) -> int | None:
    if cell_val is None:
        return None
    try:
        f = float(cell_val)
        i = int(f)
        if i == f:
            return i
    except (TypeError, ValueError):
        pass
    try:
        s = str(cell_val).strip()
        if not s:
            return None
        return int(float(s)) if "." in s else int(s)
    except (TypeError, ValueError):
        return None


def _find_submission_log_rows_for_question(ws, section_name: str, qnum_1based: int) -> list[int]:
    """1-based row indices matching Section + Q# (duplicates can exist from older appends)."""
    out: list[int] = []
    for r in range(2, (ws.max_row or 1) + 1):
        sec = ws.cell(r, 2).value
        if sec is None or str(sec).strip() != section_name:
            continue
        qn = _coerce_qnum_1based(ws.cell(r, 3).value)
        if qn == qnum_1based:
            out.append(r)
    return out


def _dedup_extra_log_rows(ws, keep_row: int, all_rows: list[int]) -> None:
    for r in sorted((x for x in all_rows if x != keep_row), reverse=True):
        ws.delete_rows(r, 1)


def _upsert_submission_log(wb, payload: dict) -> int:
    """Update or insert one row per (Section, Q#). Returns number of rows written (inserts + updates)."""
    if not isinstance(payload, dict):
        return 0
    ws = _ensure_log_sheet(wb)
    changed = 0
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    for section_name, answers in payload.items():
        if not isinstance(section_name, str) or section_name.startswith("_"):
            continue
        if not isinstance(answers, dict):
            continue
        if section_name not in wb.sheetnames or section_name == LOG_SHEET:
            continue
        s_ws = wb[section_name]
        qrows = _question_data_rows(s_ws)
        for qi_s, value in answers.items():
            try:
                qi = int(qi_s)
            except (TypeError, ValueError):
                continue
            if qi < 0 or qi >= len(qrows):
                continue
            srow = qrows[qi]
            qtext = s_ws.cell(row=srow, column=1).value
            if qtext is not None and str(qtext).strip() == "—":
                continue
            qtext = "" if qtext is None else str(qtext)
            ans = "" if value is None else str(value)
            qn = qi + 1
            matches = _find_submission_log_rows_for_question(ws, section_name, qn)
            if len(matches) > 1:
                keep = min(matches)
                _dedup_extra_log_rows(ws, keep, matches)
                matches = [keep]
            if len(matches) == 1:
                found = matches[0]
                ws.cell(row=found, column=1, value=ts)
                ws.cell(row=found, column=2, value=section_name)
                ws.cell(row=found, column=3, value=qn)
                ws.cell(row=found, column=4, value=qtext)
                ws.cell(row=found, column=5, value=ans)
                changed += 1
            else:
                if not str(ans).strip():
                    continue
                next_row = (ws.max_row or 1) + 1
                ws.cell(row=next_row, column=1, value=ts)
                ws.cell(row=next_row, column=2, value=section_name)
                ws.cell(row=next_row, column=3, value=qn)
                ws.cell(row=next_row, column=4, value=qtext)
                ws.cell(row=next_row, column=5, value=ans)
                changed += 1
    return changed


def _load_workbook_from_bytes(b: bytes):
    if not b:
        raise ValueError("empty workbook bytes")
    return load_workbook(BytesIO(b), read_only=False, data_only=False)


def _workbook_to_bytes(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def github_mode() -> bool:
    return bool(GITHUB_TOKEN and GITHUB_REPO)


@app.after_request
def _cors(resp):
    allow = os.environ.get("CORS_ALLOW_ORIGIN", "*")
    resp.headers["Access-Control-Allow-Origin"] = allow
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Submit-Token"
    return resp


@app.route("/api/executive-inputs", methods=["OPTIONS"])
def options_exec():
    return ("", 204)


@app.get("/api/health")
def health_check():
    """Readiness: GitHub write mode is active when both token and repo are set in the environment."""
    return jsonify(
        {
            "ok": True,
            "github_capture": {
                "configured": github_mode(),
                "repo": GITHUB_REPO or None,
                "file": GITHUB_FILE,
                "branch": GITHUB_BRANCH,
            },
        }
    )


@app.post("/api/executive-inputs")
def post_exec():
    if SUBMIT_TOKEN:
        if request.headers.get("X-Submit-Token", "") != SUBMIT_TOKEN:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
    if not request.is_json:
        return jsonify({"ok": False, "error": "expected application/json"}), 400
    data = request.get_json()
    if not isinstance(data, dict):
        return jsonify({"ok": False, "error": "body must be a JSON object"}), 400
    if github_mode() and (not GITHUB_FILE.endswith(".xlsx")):
        return jsonify({"ok": False, "error": "invalid GITHUB_FILE"}), 500
    try:
        rows, notice = _process_payload_with_meta(data)
    except FileNotFoundError as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    except RuntimeError as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    if rows == 0 and notice:
        return jsonify({"ok": True, "rows": 0, "message": notice, "storage": "github" if github_mode() else "local"})
    return jsonify({"ok": True, "rows": rows, "storage": "github" if github_mode() else "local"})


def _process_payload_with_meta(payload: dict) -> tuple[int, str | None]:
    if github_mode():
        for _attempt in range(8):
            state = github_get_file()
            if not state.content or not state.sha:
                raise FileNotFoundError(
                    f"Workbook {GITHUB_FILE} is missing on {GITHUB_BRANCH}. "
                    "Commit docs/Merrakii_Business_Capture.xlsx from this repository first."
                )
            wb = _load_workbook_from_bytes(state.content)
            n_changed = _upsert_submission_log(wb, payload)
            if n_changed == 0:
                return 0, "nothing to update"
            new_bytes = _workbook_to_bytes(wb)
            msg = f"chore(executive-inputs): upsert submission {datetime.now(timezone.utc).isoformat()}"
            if github_put_file(new_bytes, msg, state.sha):
                return n_changed, None
        raise RuntimeError("Could not update GitHub file (repeated conflict). Try again in a few seconds.")
    path = Path(os.environ.get("EXECUTIVE_INPUTS_XLSX", str(DEFAULT_XLSX)))
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")
    wb = load_workbook(path, read_only=False, data_only=False)
    n_changed = _upsert_submission_log(wb, payload)
    if n_changed == 0:
        return 0, "nothing to update"
    wb.save(path)
    return n_changed, None


@app.get("/")
def index():
    return send_from_directory(DELIVERABLE, "digi.html")


@app.get("/<path:filename>")
def static_docs(filename: str):
    if ".." in filename or filename.startswith("/"):
        return ("Not found", 404)

    # URL …/docs/<path> ↔ files under repository docs/ (Executive Inputs JS, etc.)
    if filename.startswith("docs/"):
        rel_within_docs = filename[5:]
        if not rel_within_docs or ".." in rel_within_docs:
            return ("Not found", 404)
        doc_target = (DOCS / rel_within_docs).resolve()
        try:
            doc_target.relative_to(DOCS.resolve())
        except ValueError:
            return ("Not found", 404)
        if doc_target.is_file():
            return send_from_directory(DOCS, rel_within_docs)
        return ("Not found", 404)

    for base in (DELIVERABLE, DOCS):
        target = (base / filename).resolve()
        try:
            target.relative_to(base.resolve())
        except ValueError:
            continue
        if target.is_file():
            return send_from_directory(base, filename)
    return ("Not found", 404)


def main():
    port = int(os.environ.get("PORT", "5050"))
    app.run(host=os.environ.get("HOST", "127.0.0.1"), port=port, debug=False, threaded=True)


if __name__ == "__main__":
    main()
